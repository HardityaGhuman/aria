"""
rag/loaders.py
--------------
File loading and text extraction. Supports PDF (per-page), Markdown, and plain
text. Markdown/text files may carry a leading YAML-style frontmatter block that
declares the document's ``department`` and ``access_tier``; this module parses
that block, strips it from the body, and stamps the resolved values onto every
returned ``Document`` so the indexer can persist them as Chroma metadata.
"""
import csv
import hashlib
import os
import re

# pyrefly: ignore [missing-import]
from langchain_core.documents import Document

SUPPORTED_EXTENSIONS = {".pdf", ".md", ".txt", ".csv", ".xlsx"}

# Default tier for any document that does not declare one. "all" means every
# authenticated user may retrieve it; the only other tier today is "hr_only".
DEFAULT_ACCESS_TIER = "all"
DEFAULT_REGION = "global"
DEFAULT_DOC_TYPE = "policy"
DEFAULT_VERSION = "2026.1"
DEFAULT_EFFECTIVE_DATE = "2026-01-01"
DEFAULT_STATUS = "active"


def file_hash(filepath: str) -> str:
    """SHA-256 of a file's bytes, used to detect changed documents on reindex."""
    digest = hashlib.sha256()
    with open(filepath, "rb") as f:
        for block in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _parse_frontmatter(text: str) -> tuple[dict, str]:
    """Split a leading ``---`` frontmatter block from the body.

    Returns ``(metadata, body)``. The parser is deliberately a flat
    ``key: value`` reader rather than a full YAML dependency — the schema is
    flat (department, access_tier, title) by design. If the file does not start
    with ``---`` or the block never closes, the whole text is treated as body
    with no metadata, so a malformed block can never break loading.
    """
    if not text.startswith("---"):
        return {}, text

    lines = text.splitlines()
    closing = next(
        (i for i in range(1, len(lines)) if lines[i].strip() == "---"),
        None,
    )
    if closing is None:
        return {}, text

    metadata: dict = {}
    for line in lines[1:closing]:
        key, sep, value = line.partition(":")
        key = key.strip()
        if not sep or not key:
            continue
        # Drop surrounding quotes and any trailing "# inline comment".
        value = re.sub(r"\s+#.*$", "", value.strip()).strip().strip("\"'")
        metadata[key] = value

    body = "\n".join(lines[closing + 1:]).lstrip("\n")
    return metadata, body


def _load_pdf_file(filepath: str) -> tuple[list[Document], dict]:
    from pypdf import PdfReader

    reader = PdfReader(filepath)
    pages = []
    filename = os.path.basename(filepath)
    for page_number, page in enumerate(reader.pages, start=1):
        page_text = page.extract_text() or ""
        if page_text.strip():
            pages.append(
                Document(
                    page_content=f"Page {page_number}\n{page_text}",
                    metadata={"page": page_number, "source": filename},
                )
            )
    # PDFs carry no frontmatter; metadata is resolved from the folder fallback.
    return pages, {}


def _load_text_like(filepath: str) -> tuple[list[Document], dict]:
    """Load a Markdown or plain-text file as a single ``Document``.

    Frontmatter is parsed and stripped so the ``---`` block is never embedded or
    retrieved. The whole document is one ``Document``; structure-aware chunking
    downstream splits it on headings.
    """
    with open(filepath, encoding="utf-8") as f:
        raw = f.read()

    frontmatter, body = _parse_frontmatter(raw)
    body = clean_text(body)
    if not body:
        return [], frontmatter

    filename = os.path.basename(filepath)
    return [Document(page_content=body, metadata={"source": filename})], frontmatter


def _read_sidecar(filepath: str) -> dict:
    """Read a companion ``<name>.meta.yaml`` flat key:value sidecar for tabular
    files (which cannot carry inline frontmatter). Reuses the frontmatter parser
    by wrapping the sidecar body in ``---`` fences. Missing sidecar -> {}."""
    sidecar = filepath + ".meta.yaml"
    if not os.path.exists(sidecar):
        return {}
    with open(sidecar, encoding="utf-8") as f:
        meta, _ = _parse_frontmatter("---\n" + f.read().strip() + "\n---\n")
    return meta


def _rows_to_documents(filename: str, header: list[str], rows: list[list[str]]) -> list[Document]:
    """Serialize each data row to a labeled key:value block — one Document per row
    so a row's fields never split across chunks (the rows-as-chunks contract)."""
    documents = []
    for row in rows:
        cells = list(row) + [""] * (len(header) - len(row))
        label_parts = [filename] + [str(cells[i]) for i in range(min(2, len(header))) if str(cells[i]).strip()]
        label = "[Table: " + " | ".join(label_parts) + "]"
        body = "\n".join(f"{col}: {cells[i]}" for i, col in enumerate(header) if str(cells[i]).strip())
        documents.append(Document(page_content=f"{label}\n{body}", metadata={"source": filename}))
    return documents


def _load_csv_file(filepath: str) -> tuple[list[Document], dict]:
    filename = os.path.basename(filepath)
    with open(filepath, newline="", encoding="utf-8") as f:
        reader = list(csv.reader(f))
    if len(reader) < 2:
        return [], _read_sidecar(filepath)
    header, rows = reader[0], reader[1:]
    return _rows_to_documents(filename, header, rows), _read_sidecar(filepath)


def _load_xlsx_file(filepath: str) -> tuple[list[Document], dict]:
    from openpyxl import load_workbook

    filename = os.path.basename(filepath)
    sheet = load_workbook(filepath, read_only=True, data_only=True).active
    grid = [[("" if c is None else str(c)) for c in r] for r in sheet.iter_rows(values_only=True)]
    if len(grid) < 2:
        return [], _read_sidecar(filepath)
    return _rows_to_documents(filename, grid[0], grid[1:]), _read_sidecar(filepath)


def load_document(filepath: str, department_fallback: str = "") -> list[Document]:
    """Load a file into ``Document`` objects, stamped with department + tier.

    ``department_fallback`` is the parent-folder name supplied by the indexer; it
    is used when a document does not declare its own ``department`` in
    frontmatter (always the case for PDFs).
    """
    extension = os.path.splitext(filepath)[1].lower()
    if extension == ".pdf":
        docs, frontmatter = _load_pdf_file(filepath)
        frontmatter = {**_read_sidecar(filepath), **frontmatter}
    elif extension in {".md", ".txt"}:
        docs, frontmatter = _load_text_like(filepath)
        # Inline frontmatter wins, but a portal-written sidecar fills any gap so
        # md/txt uploaded without a frontmatter block still pick up their tier.
        frontmatter = {**_read_sidecar(filepath), **frontmatter}
    elif extension == ".csv":
        docs, frontmatter = _load_csv_file(filepath)
    elif extension == ".xlsx":
        docs, frontmatter = _load_xlsx_file(filepath)
    else:
        return []

    department = frontmatter.get("department") or department_fallback
    access_tier = frontmatter.get("access_tier") or DEFAULT_ACCESS_TIER
    region = frontmatter.get("region") or DEFAULT_REGION
    doc_type = frontmatter.get("doc_type") or DEFAULT_DOC_TYPE
    version = frontmatter.get("version") or DEFAULT_VERSION
    effective_date = frontmatter.get("effective_date") or DEFAULT_EFFECTIVE_DATE
    status = frontmatter.get("status") or DEFAULT_STATUS
    for doc in docs:
        doc.metadata["department"] = department
        doc.metadata["access_tier"] = access_tier
        doc.metadata["region"] = region
        doc.metadata["doc_type"] = doc_type
        doc.metadata["version"] = version
        doc.metadata["effective_date"] = effective_date
        doc.metadata["status"] = status

    if extension in {".csv", ".xlsx"}:
        for doc in docs:
            doc.metadata["is_tabular"] = True
            doc.metadata["content_type"] = "reference_table"

    return docs
